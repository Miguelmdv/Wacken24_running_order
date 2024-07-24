import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from file_paths import RUN_ORD_MERGED_FILE, EXCEL_FILE
from apps.conf_app.config_env import load_env_vars

def conditional(ws, col_name, df):
    
    # Obtener los encabezados de las columnas
    columnas = {cell.value: cell.column_letter for cell in ws[1]}

    # Seleccionar la columna por nombre
    columna = columnas[col_name]
    
    # Definir el rango de celdas donde se aplicará el formato condicional
    rango_celdas = f'{columna}2:{columna}{len(df) + 1}'

    # Crear la regla de escala de color
    color_scale_rule = ColorScaleRule(
        start_type='min', start_color='F8696B',  # Rojo para el valor mínimo
        mid_type='percentile', mid_value=1, mid_color='FFEB84',  # Amarillo para el valor medio
        end_type='max', end_color='63BE7B'  # Verde para el valor máximo
    )

    # Aplicar la regla de escala de color a las celdas especificadas
    ws.conditional_formatting.add(rango_celdas, color_scale_rule)
    
    return ws


def run():
    # Cargar el archivo CSV en un DataFrame
    df = pd.read_csv(RUN_ORD_MERGED_FILE)

    # Guardar el DataFrame en un archivo Excel
    df.to_excel(EXCEL_FILE, index=False)

    # Cargar el archivo Excel para darle formato de tabla
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Crear una tabla en la hoja de trabajo
    tab = Table(displayName="Table1", ref=f"A1:{ws.dimensions.split(':')[1]}")

    # Añadir estilo a la tabla
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style


    # Añadir la tabla a la hoja de cálculo
    ws.add_table(tab)
    
    # Ajustar el ancho de las columnas
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length
        
    env_vars = load_env_vars()
    for person in range(int(env_vars["NUM_PEOPLE"])):
        ws = conditional(ws, env_vars[f"PERSON_{person+1}"], df)

    # Guardar el archivo Excel con el formato de tabla
    wb.save(EXCEL_FILE)

    print(f"\nEl archivo '{RUN_ORD_MERGED_FILE.name}' ha sido convertido a Excel y guardado como '{EXCEL_FILE}'.")

